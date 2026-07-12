"""Eksport logow Windows z opcjami V2, EVTX i raportami wieloformatowymi."""

from __future__ import annotations

import csv
import datetime as dt
import json
import os
import subprocess
import xml.etree.ElementTree as ET
from pathlib import Path

from .. import reports, winapi
from ..config import EXPORTS_DIR
from .hashing import sha256_file


DEFAULT_LOGS = [("Aplikacja", "Application"), ("Zabezpieczenia", "Security"), ("Ustawienia", "Setup"), ("System", "System")]
FIELDS = ["LogName", "TimeCreated", "Id", "Level", "Provider", "Computer", "RecordId", "ProcessId", "ThreadId", "Message"]


def _hidden_run(args: list[str], timeout: int = 900) -> subprocess.CompletedProcess:
    return subprocess.run(args, capture_output=True, text=True, encoding="mbcs", errors="replace", timeout=timeout, creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))


def _parse_datetime(value: str | None, end: bool = False) -> dt.datetime | None:
    if not value: return None
    text=str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            parsed=dt.datetime.strptime(text,fmt)
            if fmt=="%Y-%m-%d" and end: parsed=parsed.replace(hour=23,minute=59,second=59)
            return parsed
        except ValueError: pass
    raise ValueError(f"Nieprawidłowa data: {value}. Użyj RRRR-MM-DD lub RRRR-MM-DD HH:MM.")


def _query(options: dict) -> str:
    range_name=options.get("range","7d")
    if range_name in {"24h","7d","30d"}:
        milliseconds={"24h":86400000,"7d":604800000,"30d":2592000000}[range_name]
        return f"*[System[TimeCreated[timediff(@SystemTime) <= {milliseconds}]]]"
    if range_name=="custom":
        start=_parse_datetime(options.get("start")); end=_parse_datetime(options.get("end"),True)
        if start and end and start>end:raise ValueError("Data początkowa nie może być późniejsza od końcowej.")
        clauses=[]
        if start: clauses.append(f"@SystemTime >= '{start.astimezone().astimezone(dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.000Z')}'")
        if end: clauses.append(f"@SystemTime <= '{end.astimezone().astimezone(dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.999Z')}'")
        return f"*[System[TimeCreated[{' and '.join(clauses)}]]]" if clauses else "*"
    return "*"


def _events_from_xml(xml_text: str, log_name: str) -> list[dict]:
    content=xml_text.strip().lstrip("\ufeff")
    if not content:return []
    try: root=ET.fromstring(f"<Events>{content}</Events>")
    except ET.ParseError:return [{"LogName":log_name,"Message":"Nie udało się sparsować XML zwróconego przez wevtutil."}]
    records=[]; ns={"e":"http://schemas.microsoft.com/win/2004/08/events/event"}
    for event in root:
        system=event.find("e:System",ns); rendering=event.find("e:RenderingInfo",ns)
        def text(path,base=system):
            node=base.find(path,ns) if base is not None else None
            return node.text if node is not None and node.text else ""
        provider=system.find("e:Provider",ns) if system is not None else None
        time_node=system.find("e:TimeCreated",ns) if system is not None else None
        execution=system.find("e:Execution",ns) if system is not None else None
        message=text("e:Message",rendering)
        if not message:
            data=[node.text or "" for node in event.findall(".//e:EventData/e:Data",ns)]
            message=" | ".join(part for part in data if part)
        records.append({"LogName":log_name,"TimeCreated":time_node.get("SystemTime","") if time_node is not None else "","Id":text("e:EventID"),"Level":text("e:Level"),"Provider":provider.get("Name","") if provider is not None else "","Computer":text("e:Computer"),"RecordId":text("e:EventRecordID"),"ProcessId":execution.get("ProcessID","") if execution is not None else "","ThreadId":execution.get("ThreadID","") if execution is not None else "","Message":message})
    return records


def _export_channel(label: str, channel: str, query: str, limit: int, output: Path, full: bool) -> tuple[list[dict], dict]:
    xml_result=_hidden_run(["wevtutil","qe",channel,f"/q:{query}","/f:xml",f"/c:{limit}","/rd:true"])
    records=_events_from_xml(xml_result.stdout or "",channel) if xml_result.returncode==0 else [{"LogName":channel,"Message":xml_result.stderr or xml_result.stdout or f"Kod {xml_result.returncode}"}]
    evtx={"label":label,"channel":channel,"status":"POMINIĘTO","path":"","sha256":"","records":len(records),"message":"Tryb szybki"}
    if full:
        evtx_path=output/f"{label}.evtx"; export=_hidden_run(["wevtutil","epl",channel,str(evtx_path),f"/q:{query}","/ow:true"])
        success=export.returncode==0 and evtx_path.exists()
        evtx={"label":label,"channel":channel,"status":"OK" if success else "BŁĄD","path":str(evtx_path) if success else "","sha256":sha256_file(evtx_path) if success else "","records":len(records),"message":"" if success else (export.stderr or export.stdout)}
    csv_path=output/f"{label}.csv"
    with csv_path.open("w",encoding="utf-8-sig",newline="") as handle:
        writer=csv.DictWriter(handle,fieldnames=FIELDS,extrasaction="ignore"); writer.writeheader(); writer.writerows(records)
    return records,evtx


def _reports(output: Path, options: dict, sheets: dict[str,list[dict]], evtx: list[dict]) -> dict:
    stamp=dt.datetime.now().strftime("%Y%m%d_%H%M%S"); all_records=[record for records in sheets.values() for record in records]
    json_path=output/f"logi_windows_{stamp}.json"; txt_path=output/f"logi_windows_{stamp}.txt"; xlsx_path=output/f"logi_windows_{stamp}.xlsx"; pdf_path=output/f"logi_windows_{stamp}.pdf"
    json_path.write_text(json.dumps({"options":options,"events":all_records,"evtx":evtx},ensure_ascii=False,indent=2,default=str),encoding="utf-8")
    lines=["EVIDLOCK LIGHT - LOGI SYSTEMOWE WINDOWS","="*52,f"Data: {dt.datetime.now():%Y-%m-%d %H:%M:%S}",f"Tryb: {options.get('mode')}",f"Zakres: {options.get('range')}",f"Łącznie rekordów: {len(all_records)}",""]
    lines.extend(f"- {name}: {len(records)} rekordów" for name,records in sheets.items()); lines.append(""); lines.extend(f"EVTX {item['label']}: {item['status']} | {item.get('sha256') or item.get('message')}" for item in evtx)
    txt_path.write_text("\n".join(lines),encoding="utf-8")
    from openpyxl import Workbook
    workbook=Workbook(); workbook.remove(workbook.active)
    for name,records in sheets.items():
        sheet=workbook.create_sheet(name[:31]); sheet.append(FIELDS)
        for record in records:sheet.append([record.get(field,"") for field in FIELDS])
        sheet.freeze_panes="A2"; sheet.auto_filter.ref=sheet.dimensions
    summary=workbook.create_sheet("Podsumowanie",0); summary.append(["Dziennik","Rekordy","EVTX status","EVTX SHA-256"])
    for label,_channel in options.get("logs",[]):
        item=next((entry for entry in evtx if entry["label"]==label),{}); summary.append([label,len(sheets.get(label,[])),item.get("status"),item.get("sha256")])
    workbook.save(xlsx_path)
    sections=[{
        "title":"Podsumowanie eksportu",
        "rows":[
            ("Tryb",str(options.get("mode"))),
            ("Zakres czasu",str(options.get("range"))),
            ("Sortowanie",str(options.get("sort"))),
            ("Limit na dziennik",str(options.get("limit"))),
            ("Łączna liczba zdarzeń",str(len(all_records))),
        ],
    },{
        "title":"Pliki EVTX i integralność",
        "table":[[item.get("label"),item.get("channel"),item.get("status"),item.get("records"),item.get("sha256") or item.get("message") or "-"] for item in evtx],
        "headers":["Dziennik","Kanał","Status","Rekordy","SHA-256 / informacja"],
        "widths":[75,85,55,50,300],
        "wide":True,
    }]
    for name,records in sheets.items():
        sample=records[:25]
        sections.append({
            "title":f"Dziennik {name} - {len(records)} zdarzeń",
            "text":"Poniżej pokazano maksymalnie 25 najnowszych rekordów. Pełne dane znajdują się w CSV, XLSX i JSON.",
            "table":[[record.get("TimeCreated"),record.get("Id"),record.get("Level"),record.get("Provider"),record.get("Message")] for record in sample],
            "headers":["Czas","ID","Poziom","Dostawca","Komunikat"],
            "widths":[110,45,45,120,245],
            "wide":True,
        })
    reports.write_professional_pdf(
        "Eksport logów systemowych Windows",
        sections,
        pdf_path,
        subtitle="Raport zdarzeń, stanu eksportu EVTX i integralności plików wynikowych.",
        metadata={"Komputer":os.environ.get("COMPUTERNAME","Brak danych"),"Administrator":"Tak" if winapi.is_admin() else "Nie","Liczba kanałów":len(sheets)},
    )
    checksum=output/f"logi_windows_sumy_sha256_{stamp}.txt"; files=[json_path,txt_path,xlsx_path,pdf_path,*[Path(item["path"]) for item in evtx if item.get("path")]]
    checksum.write_text("\n".join(f"{sha256_file(path)}  {path.name}" for path in files if path.exists()),encoding="ascii")
    for path in [*files,checksum,*output.glob("*.csv")]:
        try:winapi.set_readonly(path,True)
        except Exception:pass
    return {"json":str(json_path),"txt":str(txt_path),"xlsx":str(xlsx_path),"pdf":str(pdf_path),"checksums":str(checksum)}


def export_logs(options: dict | None = None, output_dir: str | Path | None = None, callback=None) -> dict:
    if os.name!="nt":raise RuntimeError("Eksport logów jest dostępny tylko w Windows.")
    options=dict(options or {}); options.setdefault("mode","full"); options.setdefault("range","7d"); options.setdefault("limit",5000); options.setdefault("sort","time_desc"); options.setdefault("logs",DEFAULT_LOGS)
    base_output=Path(output_dir) if output_dir else EXPORTS_DIR/"WindowsLogs"; output=base_output/f"logs_{dt.datetime.now():%Y%m%d_%H%M%S}"; output.mkdir(parents=True,exist_ok=True)
    query=_query(options); logs=options["logs"]; sheets={}; evtx=[]; total=max(1,len(logs))
    for index,(label,channel) in enumerate(logs,1):
        if callback:callback(5+(index-1)/total*75,f"Eksport {label} ({channel})...",label,10)
        records,evtx_item=_export_channel(label,channel,query,int(options["limit"]),output,options["mode"]=="full")
        reverse=options.get("sort")=="time_desc"
        if options.get("sort") in {"time_desc","time_asc"}:records.sort(key=lambda item:item.get("TimeCreated",""),reverse=reverse)
        elif options.get("sort")=="id":records.sort(key=lambda item:int(item.get("Id") or 0) if str(item.get("Id") or "").isdigit() else 0)
        sheets[label]=records;evtx.append(evtx_item)
        if callback:callback(5+index/total*75,f"Zakończono {label}: {len(records)} rekordów",label,100)
    if callback:callback(85,"Generowanie XLSX, PDF, TXT, JSON i sum SHA-256...")
    report_files=_reports(output,options,sheets,evtx)
    result={"ok":all(item["status"] in {"OK","POMINIĘTO"} for item in evtx),"output_dir":str(output),"admin":winapi.is_admin(),"event_count":sum(len(records) for records in sheets.values()),"logs":{name:len(records) for name,records in sheets.items()},"evtx":evtx,**report_files}
    if callback:callback(100,"Eksport logów Windows zakończony")
    return result
